#!/usr/bin/env python3
"""
ComplianceOS Excel to Markdown Sync Script
This script reads the Excel file and updates the markdown file automatically.
"""

import openpyxl
import re
from datetime import datetime

def read_excel_data(file_path):
    """Read all data from the Excel file"""
    wb = openpyxl.load_workbook(file_path)
    data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = []
        
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):  # Skip empty rows
                sheet_data.append([str(cell) if cell is not None else '' for cell in row])
        
        data[sheet_name] = sheet_data
    
    return data

def generate_markdown_from_excel(data):
    """Generate markdown content from Excel data"""
    markdown_content = f"""# ComplianceOS Complete Specification
## Everything You Need to Know About Your System

---

## 📋 **DOCUMENT OVERVIEW**

This is your complete ComplianceOS reference guide containing:
- **Every button and action** on every page
- **Complete workflow mappings** 
- **Security and maintenance details**
- **JSON format** for easy modification
- **Interactive element specifications**

**📊 EXCEL WORKBOOK**: [COMPLIANCEOS_SPECIFICATION.xlsx](./COMPLIANCEOS_SPECIFICATION.xlsx) - Click to open and edit all data
**🔄 Last Updated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

---

## 🎯 **INTERACTIVE ELEMENT MAPPING**

### **Format Explanation:**
- **Page Intent** → What the page is designed to do
- **Button/Action** → Every clickable element
- **Trigger Event** → What happens when clicked
- **Action** → What the system does
- **Workflow** → n8n workflow details
- **Return Value** → What gets returned/sent
- **System** → Where it goes (email, database, etc.)

---

"""

    # Page sections mapping
    page_sections = {
        'Dashboard_Main': ('🏠 **DASHBOARD SECTION**', 'Dashboard Main (`/dashboard`)'),
        'Dashboard_Overview': ('', 'Dashboard Overview (`/dashboard/overview`)'),
        'Dashboard_KPI': ('', 'Dashboard KPI (`/dashboard/kpi`)'),
        'Dashboard_Quick_Add': ('', 'Dashboard Quick Add (`/dashboard/quick-add`)'),
        'Employees_Main': ('👥 **EMPLOYEES SECTION**', 'Employees Main (`/employees`)'),
        'Risk_Management': ('⚠️ **RISK MANAGEMENT SECTION**', 'Risk Management (`/risk`)'),
        'Equipment': ('🔧 **EQUIPMENT & CALIBRATION SECTION**', 'Equipment (`/equipment`)'),
        'Calibration': ('', 'Calibration (`/calibration`)'),
        'Work_Progress': ('📊 **WORK PROGRESS SECTION**', 'Work Progress (`/work-progress`)'),
        'OHS_Dashboard': ('🛡️ **OHS SECTION**', 'OHS Dashboard (`/ohs/dashboard`)'),
        'OHS_Actions': ('', 'OHS Actions (`/ohs/actions`)'),
        'OHS_Contractors': ('', 'OHS Contractors (`/ohs/contractors`)'),
        'Contract_Review': ('📄 **CONTRACT REVIEW SECTION**', 'Contract Review (`/contract-review`)'),
        'Training': ('🎓 **TRAINING SECTION**', 'Training (`/training`)'),
        'Audits': ('📋 **AUDITS SECTION**', 'Audits (`/audits`)'),
        'Integrations': ('🔗 **INTEGRATIONS SECTION**', 'Integrations (`/admin/integrations`)'),
        'Communication': ('📞 **SUPPORT SECTION**', 'Communication (`/support/communication`)'),
        'Reports': ('📊 **REPORTS SECTION**', 'Reports (`/reports`)'),
    }

    # Process each sheet
    for sheet_name, sheet_data in data.items():
        if sheet_name in page_sections:
            section_header, page_title = page_sections[sheet_name]
            
            if section_header:
                markdown_content += f"\n{section_header}\n\n"
            
            markdown_content += f"### {page_title}\n"
            markdown_content += f"**📊 Excel Tab**: [{sheet_name}](./COMPLIANCEOS_SPECIFICATION.xlsx#{sheet_name})\n\n"
            
            if len(sheet_data) > 1:
                # Create table
                headers = sheet_data[0]
                markdown_content += "| " + " | ".join(headers) + " |\n"
                markdown_content += "|" + "|".join([":---"] * len(headers)) + "|\n"
                
                for row in sheet_data[1:]:
                    if any(cell.strip() for cell in row):  # Skip empty rows
                        markdown_content += "| " + " | ".join(row) + " |\n"
                
                markdown_content += "\n"
        
        elif sheet_name == 'N8N_Workflows':
            markdown_content += "## 🔄 **N8N WORKFLOW DETAILS**\n"
            markdown_content += f"**📊 Excel Tab**: [{sheet_name}](./COMPLIANCEOS_SPECIFICATION.xlsx#{sheet_name})\n\n"
            markdown_content += "### **Workflow Mapping with Direct Links:**\n\n"
            
            if len(sheet_data) > 1:
                headers = sheet_data[0]
                markdown_content += "| " + " | ".join(headers) + " |\n"
                markdown_content += "|" + "|".join([":---"] * len(headers)) + "|\n"
                
                for row in sheet_data[1:]:
                    if any(cell.strip() for cell in row):
                        # Convert URLs to clickable links in markdown
                        formatted_row = []
                        for i, cell in enumerate(row):
                            if i == 5 and cell.startswith('https://'):  # URL column
                                formatted_row.append(f"[🔗 Open]({cell})")
                            else:
                                formatted_row.append(cell)
                        markdown_content += "| " + " | ".join(formatted_row) + " |\n"
                
                markdown_content += "\n"
                
                # Add quick access section
                markdown_content += "### **🚀 Quick Access Links:**\n\n"
                for row in sheet_data[1:]:
                    if len(row) >= 6 and row[5] and row[5].startswith('https://'):
                        workflow_name = row[0]
                        workflow_url = row[5]
                        workflow_id = row[6] if len(row) > 6 else 'N/A'
                        status = row[7] if len(row) > 7 else 'Unknown'
                        markdown_content += f"- **{workflow_name}** ({status}): [🔗 Open Workflow]({workflow_url}) | ID: `{workflow_id}`\n"
                
                markdown_content += "\n"
        
        elif sheet_name in ['Security', 'Data_Protection', 'Audit_Monitoring', 'Maintenance']:
            section_name = sheet_name.replace('_', ' ').title()
            markdown_content += f"## 🔒 **{section_name.upper()}**\n"
            markdown_content += f"**📊 Excel Tab**: [{sheet_name}](./COMPLIANCEOS_SPECIFICATION.xlsx#{sheet_name})\n\n"
            
            if len(sheet_data) > 1:
                headers = sheet_data[0]
                markdown_content += "| " + " | ".join(headers) + " |\n"
                markdown_content += "|" + "|".join([":---"] * len(headers)) + "|\n"
                
                for row in sheet_data[1:]:
                    if any(cell.strip() for cell in row):
                        markdown_content += "| " + " | ".join(row) + " |\n"
                
                markdown_content += "\n"
        
        elif sheet_name == 'JSON_Summary':
            markdown_content += "## 📊 **JSON FORMAT SUMMARY**\n"
            markdown_content += f"**📊 Excel Tab**: [{sheet_name}](./COMPLIANCEOS_SPECIFICATION.xlsx#{sheet_name})\n\n"
            
            # Generate JSON from the data
            json_data = {}
            for row in sheet_data[1:]:
                if len(row) >= 2 and row[0] and row[1]:
                    json_data[row[0]] = row[1]
            
            markdown_content += "```json\n"
            markdown_content += "{\n"
            markdown_content += '  "complianceos_complete_specification": {\n'
            for key, value in json_data.items():
                if key and value:
                    markdown_content += f'    "{key}": "{value}",\n'
            markdown_content += "  }\n"
            markdown_content += "}\n"
            markdown_content += "```\n\n"
        
        elif sheet_name == 'Maintenance_Checklists':
            markdown_content += "## 📋 **MAINTENANCE CHECKLISTS**\n"
            markdown_content += f"**📊 Excel Tab**: [{sheet_name}](./COMPLIANCEOS_SPECIFICATION.xlsx#{sheet_name})\n\n"
            
            if len(sheet_data) > 1:
                headers = sheet_data[0]
                markdown_content += "| " + " | ".join(headers) + " |\n"
                markdown_content += "|" + "|".join([":---"] * len(headers)) + "|\n"
                
                for row in sheet_data[1:]:
                    if any(cell.strip() for cell in row):
                        markdown_content += "| " + " | ".join(row) + " |\n"
                
                markdown_content += "\n"
        
        elif sheet_name == 'Security_Contacts':
            markdown_content += "## 📞 **SECURITY CONTACTS & ESCALATION**\n"
            markdown_content += f"**📊 Excel Tab**: [{sheet_name}](./COMPLIANCEOS_SPECIFICATION.xlsx#{sheet_name})\n\n"
            
            if len(sheet_data) > 1:
                headers = sheet_data[0]
                markdown_content += "| " + " | ".join(headers) + " |\n"
                markdown_content += "|" + "|".join([":---"] * len(headers)) + "|\n"
                
                for row in sheet_data[1:]:
                    if any(cell.strip() for cell in row):
                        markdown_content += "| " + " | ".join(row) + " |\n"
                
                markdown_content += "\n"

    # Add footer
    markdown_content += """---

**This specification gives you complete visibility into every interactive element, security measure, and maintenance requirement in ComplianceOS!** 🚀

**📊 EXCEL WORKBOOK**: [COMPLIANCEOS_SPECIFICATION.xlsx](./COMPLIANCEOS_SPECIFICATION.xlsx) - Click to open and edit all data
"""

    return markdown_content

def sync_excel_to_markdown(excel_file, markdown_file):
    """Sync Excel file to markdown file"""
    print(f"🔄 Reading Excel file: {excel_file}")
    data = read_excel_data(excel_file)
    
    print(f"📝 Generating markdown content...")
    markdown_content = generate_markdown_from_excel(data)
    
    print(f"💾 Writing markdown file: {markdown_file}")
    with open(markdown_file, 'w', encoding='utf-8') as f:
        f.write(markdown_content)
    
    print(f"✅ Sync complete! Markdown file updated.")
    print(f"📊 Processed {len(data)} Excel sheets")
    print(f"🕒 Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

if __name__ == "__main__":
    excel_file = "COMPLIANCEOS_SPECIFICATION.xlsx"
    markdown_file = "COMPLIANCEOS_COMPLETE_SPECIFICATION.md"
    
    try:
        sync_excel_to_markdown(excel_file, markdown_file)
    except Exception as e:
        print(f"❌ Error: {e}")
        print("Make sure the Excel file exists and is not open in Excel.")
